# app.py
import streamlit as st
import pandas as pd
import os
from geopy.distance import geodesic
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pydeck as pdk

# ---------------------------
# Config
# ---------------------------
st.set_page_config(page_title="Site Nearest Finder (with Map)", layout="wide")
st.title("📍 Site Nearest Finder — Map & Save Mapinfo")
st.write("Upload Mapinfo, upload ATND atau input manual, generate hasil Excel + lihat peta interaktif.")

SAVE_MAPINFO_FILENAME = "saved_mapinfo.xlsx"

# ---------------------------
# Helper: read excel with proper engine based on ext
# ---------------------------
def read_excel_bytesio(uploaded_file):
    name = uploaded_file.name
    if name.lower().endswith(".xls"):
        return pd.read_excel(uploaded_file, engine="xlrd")
    else:
        return pd.read_excel(uploaded_file, engine="openpyxl")

# ---------------------------
# Load saved Mapinfo if exists on disk
# ---------------------------
if "df_map" not in st.session_state:
    st.session_state.df_map = None
    st.session_state.mapinfo_last_updated = None

if os.path.exists(SAVE_MAPINFO_FILENAME) and st.session_state.df_map is None:
    try:
        df_map_saved = pd.read_excel(SAVE_MAPINFO_FILENAME, engine="openpyxl")
        st.session_state.df_map = df_map_saved
        st.session_state.mapinfo_last_updated = datetime.fromtimestamp(os.path.getmtime(SAVE_MAPINFO_FILENAME))
    except Exception as e:
        # ignore load errors, user will reupload
        st.warning(f"Warning: gagal load saved_mapinfo: {e}")

# ---------------------------
# Mapinfo upload / update / delete
# ---------------------------
st.subheader("📂 Upload / Update Mapinfo (disimpan sementara)")
col1, col2, col3 = st.columns([6,2,2])

with col1:
    uploaded_mapinfo = st.file_uploader("Upload Mapinfo (.xls/.xlsx) — akan disimpan sebagai saved_mapinfo.xlsx", type=["xls","xlsx"], key="mapinfo_uploader")
with col2:
    if st.button("🗑 Hapus saved Mapinfo"):
        if os.path.exists(SAVE_MAPINFO_FILENAME):
            try:
                os.remove(SAVE_MAPINFO_FILENAME)
                st.session_state.df_map = None
                st.session_state.mapinfo_last_updated = None
                st.success("✅ saved_mapinfo.xlsx dihapus.")
            except Exception as e:
                st.error(f"Gagal menghapus: {e}")
        else:
            st.info("Tidak ada saved_mapinfo.xlsx untuk dihapus.")
with col3:
    if st.button("🔄 Reload saved Mapinfo dari disk"):
        if os.path.exists(SAVE_MAPINFO_FILENAME):
            try:
                st.session_state.df_map = pd.read_excel(SAVE_MAPINFO_FILENAME, engine="openpyxl")
                st.session_state.mapinfo_last_updated = datetime.fromtimestamp(os.path.getmtime(SAVE_MAPINFO_FILENAME))
                st.success("✅ Mapinfo di-reload dari saved_mapinfo.xlsx")
            except Exception as e:
                st.error(f"Gagal reload: {e}")
        else:
            st.info("Tidak ada saved_mapinfo.xlsx di disk.")

if uploaded_mapinfo:
    try:
        df_map = read_excel_bytesio(uploaded_mapinfo)
        required_cols_map = {"Site ID", "Longitude", "Latitude", "BSC"}
        if not required_cols_map.issubset(df_map.columns):
            st.error(f"Kolom wajib hilang di Mapinfo: {required_cols_map - set(df_map.columns)}")
        else:
            # normalize column names (strip)
            df_map.columns = [c.strip() for c in df_map.columns]
            st.session_state.df_map = df_map
            # save to disk as xlsx (use openpyxl engine)
            try:
                df_map.to_excel(SAVE_MAPINFO_FILENAME, index=False, engine="openpyxl")
                st.session_state.mapinfo_last_updated = datetime.fromtimestamp(os.path.getmtime(SAVE_MAPINFO_FILENAME))
                st.success(f"✅ Mapinfo disimpan ke memori dan file '{SAVE_MAPINFO_FILENAME}'. ({len(df_map)} records)")
            except Exception as e:
                st.warning(f"Mapinfo dimuat ke memori, tetapi gagal menyimpan file: {e}")
    except Exception as e:
        st.error(f"Gagal membaca Mapinfo: {e}")

# show status of mapinfo
if st.session_state.df_map is not None:
    st.info(f"Mapinfo loaded — {len(st.session_state.df_map)} rows. Last saved: {st.session_state.mapinfo_last_updated}")
else:
    st.warning("Mapinfo belum diupload. Upload dulu agar bisa mencari 3 site terdekat.")

st.divider()

# ---------------------------
# Input mode: upload ATND or manual input
# ---------------------------
st.subheader("🧭 Pilih Cara Input Site")
mode = st.radio("Mode input:", ["Upload ATND file", "Input Manual (grid)"])

df_sites = None

if mode == "Upload ATND file":
    uploaded_sites = st.file_uploader("Upload ATND (.xls/.xlsx) — file yang berisi Site ID, Longitude, Latitude", type=["xls","xlsx"], key="atnd_uploader")
    if uploaded_sites:
        try:
            df_sites_tmp = read_excel_bytesio(uploaded_sites)
            required_cols_sites = {"Site ID", "Longitude", "Latitude"}
            if not required_cols_sites.issubset(df_sites_tmp.columns):
                st.error(f"Kolom wajib hilang di ATND: {required_cols_sites - set(df_sites_tmp.columns)}")
            else:
                df_sites = df_sites_tmp[[col for col in df_sites_tmp.columns if col in df_sites_tmp.columns]]
                df_sites = df_sites.reset_index(drop=True)
                st.success(f"ATND loaded — {len(df_sites)} sites.")
        except Exception as e:
            st.error(f"Gagal membaca ATND: {e}")

else:
    st.write("Masukkan data site secara manual (bisa tambah baris).")
    # default empty dataframe columns
    default_df = pd.DataFrame(columns=["Site ID", "NE ID", "Sitename", "Longitude", "Latitude"])
    df_sites = st.data_editor(default_df, num_rows="dynamic", key="manual_sites")
    if not df_sites.empty:
        st.success(f"{len(df_sites)} site dimasukkan.")

st.divider()

# ---------------------------
# Calculation: nearest 3
# ---------------------------
def get_top3_nearest(lat, lon, df_map_local):
    # ensure lat/lon numeric
    try:
        target = (float(lat), float(lon))
    except Exception:
        return ["ERR", "ERR", "ERR"]
    df_map_local = df_map_local.copy()
    # compute distance
    df_map_local["Distance_km"] = df_map_local.apply(
        lambda r: geodesic(target, (r["Latitude"], r["Longitude"])).km, axis=1
    )
    nearest = df_map_local.nsmallest(3, "Distance_km")[["Site ID", "BSC", "Distance_km"]]
    # if less than 3 in map, pad
    out = []
    for i in range(3):
        if i < len(nearest):
            out.append(f"{nearest.iloc[i]['Site ID']} - {nearest.iloc[i]['BSC']} ({nearest.iloc[i]['Distance_km']:.2f} km)")
        else:
            out.append("")
    return out

# ---------------------------
# Generate button
# ---------------------------
map_view = None  # will store pydeck Deck if generated
if st.button("🚀 Generate Hasil & Show Map"):
    if st.session_state.df_map is None:
        st.error("Harap upload Mapinfo terlebih dahulu.")
    elif df_sites is None or df_sites.empty:
        st.error("Harap upload ATND atau isi site manual terlebih dahulu.")
    else:
        df_map_ref = st.session_state.df_map.copy()
        # validate lat/lon types in df_map_ref
        if not {"Latitude","Longitude"}.issubset(df_map_ref.columns):
            st.error("Mapinfo tidak mengandung kolom Latitude/Longitude yang benar.")
        else:
            with st.spinner("Menghitung 3 site terdekat untuk setiap site ..."):
                nearest_data = []
                # collectors for map markers and lines
                markers = []
                lines = []

                for idx, row in df_sites.iterrows():
                    lat = row.get("Latitude")
                    lon = row.get("Longitude")
                    # if missing lat/lon skip
                    try:
                        latf = float(lat)
                        lonf = float(lon)
                    except Exception:
                        nearest_data.append(["ERR","ERR","ERR"])
                        continue

                    top3 = get_top3_nearest(latf, lonf, df_map_ref)
                    nearest_data.append(top3)

                    # add marker for the input site
                    markers.append({
                        "site_id": str(row.get("Site ID", "")),
                        "sitename": str(row.get("Sitename","")),
                        "lon": lonf,
                        "lat": latf,
                        "type": "input"
                    })

                    # parse each nearest to add markers/lines
                    # nearest format: "SITEID - BSC (X.XX km)"
                    # we'll find the actual coords from df_map_ref using Site ID
                    for i in range(3):
                        if top3[i] and top3[i] != "ERR":
                            # extract site id before ' - '
                            try:
                                site_id = top3[i].split(" - ")[0].strip()
                                match_row = df_map_ref[df_map_ref["Site ID"].astype(str) == site_id]
                                if not match_row.empty:
                                    mlat = float(match_row.iloc[0]["Latitude"])
                                    mlon = float(match_row.iloc[0]["Longitude"])
                                    markers.append({
                                        "site_id": site_id,
                                        "sitename": str(match_row.iloc[0].get("BSC","")),
                                        "lon": mlon,
                                        "lat": mlat,
                                        "type": "nearest"
                                    })
                                    # add a line from input site to this nearest
                                    lines.append({
                                        "from_lon": lonf,
                                        "from_lat": latf,
                                        "to_lon": mlon,
                                        "to_lat": mlat
                                    })
                            except Exception:
                                pass

                # build result dataframe
                nearest_df = pd.DataFrame(nearest_data, columns=["Nearest Site 1","Nearest Site 2","Nearest Site 3"])
                df_result = pd.concat([df_sites.reset_index(drop=True), nearest_df], axis=1)

                # Save XLSX to buffer with formatting
                output = BytesIO()
                df_result.to_excel(output, index=False, engine="openpyxl")
                output.seek(0)
                wb = load_workbook(output)
                ws = wb.active
                ws.freeze_panes = "A2"

                header_fill = PatternFill(start_color="0B3D91", end_color="0B3D91", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                header_align = Alignment(horizontal="center", vertical="center")
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                    cell.border = border

                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        cell.border = border

                for col in ws.columns:
                    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(10, max_length + 2)

                buf = BytesIO()
                wb.save(buf)
                buf.seek(0)

                # offer download
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"ATND_Result_{timestamp}.xlsx"
                st.success("✅ Proses selesai — download hasil di bawah.")
                st.download_button(
                    "📥 Download Hasil Excel",
                    data=buf,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                # build pydeck map
                if markers:
                    # deduplicate markers by (site_id, lon, lat)
                    unique_markers = {(m["site_id"], m["lon"], m["lat"], m["type"], m["sitename"]): m for m in markers}
                    marker_list = list(unique_markers.values())

                    marker_df = pd.DataFrame(marker_list)

                    # pydeck expects [lon, lat] for positions
                    marker_df["coordinates"] = marker_df.apply(lambda r: [r["lon"], r["lat"]], axis=1)
                    # color column: input = blue, nearest = green
                    marker_df["color"] = marker_df["type"].apply(lambda t: [0,116,217] if t=="input" else [34,139,34])

                    scatter = pdk.Layer(
                        "ScatterplotLayer",
                        data=marker_df,
                        get_position="coordinates",
                        get_fill_color="color",
                        get_radius=100,
                        radius_min_pixels=5,
                        radius_max_pixels=20,
                        pickable=True
                    )

                    # lines df
                    if lines:
                        lines_df = pd.DataFrame(lines)
                        lines_df["from"] = lines_df.apply(lambda r: [r["from_lon"], r["from_lat"]], axis=1)
                        lines_df["to"] = lines_df.apply(lambda r: [r["to_lon"], r["to_lat"]], axis=1)
                        line_layer = pdk.Layer(
                            "LineLayer",
                            data=lines_df,
                            get_source_position="from",
                            get_target_position="to",
                            get_width=2,
                            get_color=[100,100,100],
                            pickable=False
                        )
                        layers = [scatter, line_layer]
                    else:
                        layers = [scatter]

                    # set view state: center on first input site or average of markers
                    center_lon = float(marker_df.iloc[0]["lon"])
                    center_lat = float(marker_df.iloc[0]["lat"])
                    view_state = pdk.ViewState(latitude=center_lat, longitude=center_lon, zoom=11, pitch=0)

                    r = pdk.Deck(layers=layers, initial_view_state=view_state, tooltip={"text":"{site_id}\n{sitename}"})
                    st.pydeck_chart(r)
                else:
                    st.info("Tidak ada marker untuk ditampilkan di peta.")

# Footer notes
st.caption("Catatan: saved_mapinfo.xlsx disimpan di filesystem app. Pada Streamlit Cloud file ini bertahan selama instance hidup — bisa hilang saat redeploy. Untuk persistence permanen gunakan Google Drive / S3 / GitHub.")
